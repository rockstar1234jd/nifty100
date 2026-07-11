"""Sprint 3 screener engine."""

from __future__ import annotations

import logging
import sqlite3
from pathlib import Path
from typing import Any

import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from src.analytics.ratios import DEFAULT_DB_PATH, DEFAULT_OUTPUT_DIR

logger = logging.getLogger(__name__)

ROOT_DIR = Path(__file__).resolve().parents[2]
DEFAULT_CONFIG_PATH = ROOT_DIR / "config" / "screener_config.yaml"

PASS_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
FAIL_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def load_screener_config(config_path: Path | None = None) -> dict[str, Any]:
	path = config_path or DEFAULT_CONFIG_PATH
	with path.open("r", encoding="utf-8") as handle:
		return yaml.safe_load(handle) or {}


def load_latest_screening_frame(db_path: Path | None = None) -> pd.DataFrame:
	path = db_path or DEFAULT_DB_PATH
	with sqlite3.connect(path) as conn:
		query = """
		SELECT
			fr.company_id,
			fr.year,
			c.company_name,
			c.roe_percentage AS return_on_equity_pct,
			COALESCE(s.broad_sector, 'Unclassified') AS broad_sector,
			fr.debt_to_equity,
			fr.free_cash_flow_cr,
			fr.revenue_cagr_5yr,
			fr.pat_cagr_5yr,
			fr.operating_profit_margin_pct,
			m.pe_ratio,
			m.pb_ratio,
			m.dividend_yield_pct,
			fr.interest_coverage,
			m.market_cap_crore,
			pl.net_profit AS net_profit,
			fr.eps_cagr_5yr,
			fr.asset_turnover,
			pl.sales,
			prev.debt_to_equity AS debt_to_equity_prev_year,
			fr.composite_quality_score,
			pg.peer_group_name,
			COALESCE(pg.is_benchmark, 0) AS is_benchmark,
			fr.icr_label
		FROM financial_ratios fr
		JOIN companies c ON c.id = fr.company_id
		LEFT JOIN sectors s ON s.company_id = fr.company_id
		LEFT JOIN market_cap m ON m.company_id = fr.company_id AND m.year = CAST(substr(fr.year, 1, 4) AS INTEGER)
		LEFT JOIN profitandloss pl ON pl.company_id = fr.company_id AND pl.year = fr.year
		LEFT JOIN financial_ratios prev ON prev.company_id = fr.company_id AND prev.year = printf('%04d-%s', CAST(substr(fr.year, 1, 4) AS INTEGER) - 1, substr(fr.year, 6))
		LEFT JOIN peer_groups pg ON pg.company_id = fr.company_id
		ORDER BY fr.company_id, fr.year
		"""
		frame = pd.read_sql_query(query, conn)
	if frame.empty:
		return frame
	frame["year"] = frame["year"].astype(str)
	frame = frame.sort_values(["company_id", "year"]).groupby("company_id", as_index=False).tail(1).reset_index(drop=True)
	return frame


def _numeric(series: pd.Series) -> pd.Series:
	return pd.to_numeric(series, errors="coerce")


def apply_filter(frame: pd.DataFrame, rule: dict[str, Any], threshold: float) -> pd.Series:
	value = _numeric(frame[rule["column"]])
	operator = rule.get("operator", ">=")
	if rule.get("sector_bypass"):
		sector_mask = frame["broad_sector"].astype(str).isin(rule["sector_bypass"])
		comparison = value <= threshold if operator == "<=" else value >= threshold
		return (sector_mask | comparison).fillna(False)
	if rule.get("icr_passthrough"):
		debt_free = frame.get("icr_label", pd.Series(index=frame.index, dtype=object)).astype(str).eq("Debt Free")
		return (debt_free | (value >= threshold)).fillna(False)
	if operator == "<=":
		return (value <= threshold).fillna(False)
	return (value >= threshold).fillna(False)


def sort_screened_frame(frame: pd.DataFrame) -> pd.DataFrame:
	if frame.empty:
		return frame.copy()
	result = frame.copy()
	result["composite_quality_score"] = _numeric(result["composite_quality_score"])
	result["sector_relative_score"] = result.groupby("broad_sector")["composite_quality_score"].transform(lambda s: s.rank(pct=True, ascending=True) * 100)
	result = result.sort_values(["composite_quality_score", "sector_relative_score", "market_cap_crore"], ascending=[False, False, False])
	return result.reset_index(drop=True)


def apply_preset(frame: pd.DataFrame, config: dict[str, Any], preset_name: str) -> pd.DataFrame:
	preset = config["screener"]["presets"][preset_name]
	filters = config["screener"]["filters"]
	filtered = frame.copy()
	for filter_name, threshold in preset["thresholds"].items():
		filtered = filtered.loc[apply_filter(filtered, filters[filter_name], threshold)]
	if preset_name == "turnaround_watch" and "debt_to_equity_prev_year" in filtered.columns:
		filtered = filtered.loc[(filtered["debt_to_equity"].fillna(0) < filtered["debt_to_equity_prev_year"].fillna(0)) | (filtered["debt_to_equity"].fillna(0) == 0)]
	return sort_screened_frame(filtered)


def run_all_presets(db_path: Path | None = None, config_path: Path | None = None) -> dict[str, pd.DataFrame]:
	config = load_screener_config(config_path)
	frame = load_latest_screening_frame(db_path)
	results: dict[str, pd.DataFrame] = {}
	for preset_name in config.get("screener", {}).get("presets", {}):
		results[preset_name] = apply_preset(frame, config, preset_name)
	return results


def write_screener_workbook(results: dict[str, pd.DataFrame], output_path: Path | None = None) -> Path:
	path = output_path or (DEFAULT_OUTPUT_DIR / "screener_output.xlsx")
	path.parent.mkdir(parents=True, exist_ok=True)
	config = load_screener_config()
	presets = config.get("screener", {}).get("presets", {})
	filters = config.get("screener", {}).get("filters", {})
	with pd.ExcelWriter(path, engine="openpyxl") as writer:
		for preset_name, frame in results.items():
			frame.to_excel(writer, sheet_name=preset_name[:31], index=False)
			worksheet = writer.book[preset_name[:31]]
			for cell in worksheet[1]:
				cell.fill = HEADER_FILL
				cell.font = HEADER_FONT
			preset = presets.get(preset_name, {})
			thresholds = preset.get("thresholds", {})
			column_map = {column_name: index + 1 for index, column_name in enumerate(frame.columns)}
			for filter_name, threshold in thresholds.items():
				rule = filters[filter_name]
				column_name = rule["column"]
				if column_name not in column_map:
					continue
				column_index = column_map[column_name]
				operator = rule.get("operator", ">=")
				for row_index in range(2, worksheet.max_row + 1):
					cell = worksheet.cell(row=row_index, column=column_index)
					value = cell.value
					if value is None:
						continue
					if operator == "<=":
						cell.fill = PASS_FILL if value <= threshold else FAIL_FILL
					else:
						cell.fill = PASS_FILL if value >= threshold else FAIL_FILL
	return path


def main() -> int:
	results = run_all_presets()
	write_screener_workbook(results)
	logger.info("Screener outputs written: %d preset(s)", len(results))
	return 0


if __name__ == "__main__":
	raise SystemExit(main())
