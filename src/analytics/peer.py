"""Sprint 3 peer comparison engine."""

from __future__ import annotations

import logging
import sqlite3
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl import load_workbook

from src.analytics.ratios import DEFAULT_DB_PATH, DEFAULT_OUTPUT_DIR

logger = logging.getLogger(__name__)

ROOT_DIR = Path(__file__).resolve().parents[2]
DEFAULT_RADAR_DIR = ROOT_DIR / "reports" / "radar_charts"

GREEN_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
RED_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")
GOLD_FILL = PatternFill(fill_type="solid", fgColor="FFD966")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MEDIAN_FILL = PatternFill(fill_type="solid", fgColor="E7E6E6")

PEER_METRICS = [
	"return_on_equity_pct",
	"return_on_capital_employed_pct",
	"operating_profit_margin_pct",
	"debt_to_equity",
	"interest_coverage",
	"asset_turnover",
	"free_cash_flow_cr",
	"revenue_cagr_5yr",
	"pat_cagr_5yr",
	"dividend_yield_pct",
]

RADAR_METRICS = [
	"return_on_equity_pct",
	"return_on_capital_employed_pct",
	"operating_profit_margin_pct",
	"debt_to_equity",
	"interest_coverage",
	"asset_turnover",
	"free_cash_flow_cr",
	"dividend_yield_pct",
]


def load_peer_base_frame(db_path: Path | None = None) -> pd.DataFrame:
	path = db_path or DEFAULT_DB_PATH
	with sqlite3.connect(path) as conn:
		query = """
		SELECT
			fr.company_id,
			c.company_name,
			c.roe_percentage AS return_on_equity_pct,
			COALESCE(s.broad_sector, 'Unclassified') AS broad_sector,
			COALESCE(pg.peer_group_name, '') AS peer_group_name,
			COALESCE(pg.is_benchmark, 0) AS is_benchmark,
			fr.year,
			fr.return_on_capital_employed_pct,
			fr.operating_profit_margin_pct,
			fr.debt_to_equity,
			fr.interest_coverage,
			fr.asset_turnover,
			fr.free_cash_flow_cr,
			fr.revenue_cagr_5yr,
			fr.pat_cagr_5yr,
			m.dividend_yield_pct,
			m.market_cap_crore,
			fr.composite_quality_score,
			fr.net_profit_margin_pct,
			fr.eps_cagr_5yr
		FROM financial_ratios fr
		JOIN companies c ON c.id = fr.company_id
		LEFT JOIN sectors s ON s.company_id = fr.company_id
		LEFT JOIN peer_groups pg ON pg.company_id = fr.company_id
		LEFT JOIN market_cap m ON m.company_id = fr.company_id AND m.year = CAST(substr(fr.year, 1, 4) AS INTEGER)
		ORDER BY fr.company_id, fr.year
		"""
		frame = pd.read_sql_query(query, conn)
	if frame.empty:
		return frame
	frame["year"] = frame["year"].astype(str)
	return frame.sort_values(["company_id", "year"]).groupby("company_id", as_index=False).tail(1).reset_index(drop=True)


def _percent_rank(series: pd.Series, invert: bool = False) -> pd.Series:
	value = pd.to_numeric(series, errors="coerce")
	if value.dropna().empty:
		return pd.Series([50.0] * len(series), index=series.index)
	if value.nunique(dropna=True) <= 1:
		return pd.Series([50.0] * len(series), index=series.index)
	rank = value.rank(pct=True, method="average") * 100
	return 100 - rank if invert else rank


def compute_peer_percentiles(frame: pd.DataFrame) -> pd.DataFrame:
	if frame.empty:
		return frame.copy()
	working = frame.copy()
	for metric in PEER_METRICS:
		if metric not in working.columns:
			working[metric] = np.nan
	for metric in PEER_METRICS:
		invert = metric == "debt_to_equity"
		working[f"{metric}_percentile"] = working.groupby("peer_group_name")[metric].transform(lambda s: _percent_rank(s, invert=invert))
	return working


def write_peer_percentiles_table(frame: pd.DataFrame, db_path: Path | None = None) -> None:
	path = db_path or DEFAULT_DB_PATH
	with sqlite3.connect(path) as conn:
		conn.execute("DROP TABLE IF EXISTS peer_percentiles")
		frame.to_sql("peer_percentiles", conn, if_exists="replace", index=False)
		conn.commit()


def _radar_labels() -> list[str]:
	return [
		"ROE",
		"ROCE",
		"OPM",
		"Low Debt",
		"ICR",
		"Asset Turnover",
		"FCF",
		"Dividend Yield",
	]


def _radar_values(row: pd.Series) -> list[float]:
	return [
		float(row.get("return_on_equity_pct_percentile", 50)),
		float(row.get("return_on_capital_employed_pct_percentile", 50)),
		float(row.get("operating_profit_margin_pct_percentile", 50)),
		float(row.get("debt_to_equity_percentile", 50)),
		float(row.get("interest_coverage_percentile", 50)),
		float(row.get("asset_turnover_percentile", 50)),
		float(row.get("free_cash_flow_cr_percentile", 50)),
		float(row.get("dividend_yield_pct_percentile", 50)),
	]


def _universe_percentile(base: pd.DataFrame, column: str, value: object, invert: bool = False) -> float:
	series = pd.to_numeric(base[column], errors="coerce")
	if series.dropna().empty or pd.isna(value):
		return 50.0
	score = float((series <= float(value)).mean() * 100)
	return 100 - score if invert else score


def _plot_radar(values: list[float], labels: list[str], title: str, output_path: Path) -> None:
	count = len(labels)
	angles = np.linspace(0, 2 * np.pi, count, endpoint=False).tolist()
	values = values + values[:1]
	angles = angles + angles[:1]
	fig = plt.figure(figsize=(9, 9))
	ax = plt.subplot(111, polar=True)
	ax.set_theta_offset(np.pi / 2)
	ax.set_theta_direction(-1)
	ax.set_thetagrids(np.degrees(angles[:-1]), labels, fontsize=11)
	ax.set_ylim(0, 100)
	ax.plot(angles, values, linewidth=2.5, color="#155EEF")
	ax.fill(angles, values, color="#155EEF", alpha=0.20)
	ax.plot(angles, [50] * len(angles), linewidth=1.5, linestyle="--", color="#555555")
	ax.set_title(title, fontsize=14, pad=24)
	output_path.parent.mkdir(parents=True, exist_ok=True)
	fig.tight_layout()
	fig.savefig(output_path, dpi=160)
	plt.close(fig)


def _apply_peer_sheet_styling(worksheet) -> None:
	for cell in worksheet[1]:
		cell.fill = HEADER_FILL
		cell.font = HEADER_FONT
	column_lookup = {cell.value: cell.column for cell in worksheet[1]}
	percentile_columns = [name for name in column_lookup if isinstance(name, str) and name.endswith("_percentile")]
	for row_index in range(2, worksheet.max_row + 1):
		is_median_row = row_index == worksheet.max_row
		is_benchmark_row = worksheet.cell(row=row_index, column=column_lookup.get("is_benchmark", 1)).value == 1
		for name in percentile_columns:
			cell = worksheet.cell(row=row_index, column=column_lookup[name])
			value = pd.to_numeric(cell.value, errors="coerce")
			if pd.isna(value):
				continue
			if value >= 75:
				cell.fill = GREEN_FILL
			elif value >= 25:
				cell.fill = YELLOW_FILL
			else:
				cell.fill = RED_FILL
		if is_benchmark_row:
			for cell in worksheet[row_index]:
				cell.fill = GOLD_FILL
		if is_median_row:
			for cell in worksheet[row_index]:
				if cell.fill.fill_type is None:
					cell.fill = MEDIAN_FILL


def write_peer_comparison_workbook(frame: pd.DataFrame, output_path: Path | None = None) -> Path:
	path = output_path or (DEFAULT_OUTPUT_DIR / "peer_comparison.xlsx")
	path.parent.mkdir(parents=True, exist_ok=True)
	with pd.ExcelWriter(path, engine="openpyxl") as writer:
		for group_name, group_frame in sorted(frame.loc[frame["peer_group_name"].astype(str) != ""].groupby("peer_group_name"), key=lambda item: item[0]):
			export_columns = ["company_id", "company_name", "broad_sector", "market_cap_crore", "is_benchmark", *PEER_METRICS, *[f"{metric}_percentile" for metric in PEER_METRICS]]
			export = group_frame[export_columns].copy()
			median_row = {column: None for column in export_columns}
			median_row["company_id"] = "MEDIAN"
			median_row["company_name"] = "Median"
			median_row["broad_sector"] = ""
			median_row["is_benchmark"] = 0
			median_row["market_cap_crore"] = group_frame["market_cap_crore"].median()
			for metric in PEER_METRICS:
				median_row[metric] = group_frame[metric].median()
				median_row[f"{metric}_percentile"] = group_frame[f"{metric}_percentile"].median()
			export = pd.concat([export, pd.DataFrame([median_row])], ignore_index=True)
			export.to_excel(writer, sheet_name=group_name[:31], index=False)
			_apply_peer_sheet_styling(writer.book[group_name[:31]])
	return path


def write_radar_charts(frame: pd.DataFrame, radar_dir: Path | None = None) -> list[Path]:
	dir_path = radar_dir or DEFAULT_RADAR_DIR
	paths: list[Path] = []
	labels = _radar_labels()
	grouped = frame.loc[frame["peer_group_name"].astype(str) != ""]
	for _, row in grouped.iterrows():
		output_path = dir_path / f"{row['company_id']}_radar.png"
		_plot_radar(_radar_values(row), labels, f"{row['company_id']} Peer Profile", output_path)
		paths.append(output_path)
	return paths


def write_standalone_radar(base_frame: pd.DataFrame, row: pd.Series, output_path: Path | None = None) -> Path:
	path = output_path or (DEFAULT_RADAR_DIR / f"{row['company_id']}_radar.png")
	labels = _radar_labels()
	row = pd.Series({
		"return_on_equity_pct_percentile": _universe_percentile(base_frame, "return_on_equity_pct", row.get("return_on_equity_pct")),
		"return_on_capital_employed_pct_percentile": _universe_percentile(base_frame, "return_on_capital_employed_pct", row.get("return_on_capital_employed_pct")),
		"operating_profit_margin_pct_percentile": _universe_percentile(base_frame, "operating_profit_margin_pct", row.get("operating_profit_margin_pct")),
		"debt_to_equity_percentile": _universe_percentile(base_frame, "debt_to_equity", row.get("debt_to_equity"), invert=True),
		"interest_coverage_percentile": _universe_percentile(base_frame, "interest_coverage", row.get("interest_coverage")),
		"asset_turnover_percentile": _universe_percentile(base_frame, "asset_turnover", row.get("asset_turnover")),
		"free_cash_flow_cr_percentile": _universe_percentile(base_frame, "free_cash_flow_cr", row.get("free_cash_flow_cr")),
		"dividend_yield_pct_percentile": _universe_percentile(base_frame, "dividend_yield_pct", row.get("dividend_yield_pct")),
	})
	_plot_radar(_radar_values(row), labels, f"{row.get('company_id', 'Company')} vs Nifty 100 Avg", path)
	return path


def run_peer_engine(db_path: Path | None = None) -> pd.DataFrame:
	base = load_peer_base_frame(db_path)
	if base.empty:
		return base
	grouped = base.loc[base["peer_group_name"].astype(str) != ""].copy()
	frame = compute_peer_percentiles(grouped)
	write_peer_percentiles_table(frame, db_path)
	return frame


def main() -> int:
	base = load_peer_base_frame()
	if base.empty:
		logger.warning("No peer data available")
		return 0
	grouped = base.loc[base["peer_group_name"].astype(str) != ""].copy()
	ungrouped = base.loc[base["peer_group_name"].astype(str) == ""].copy()
	frame = compute_peer_percentiles(grouped)
	write_peer_percentiles_table(frame)
	write_peer_comparison_workbook(frame)
	write_radar_charts(frame)
	for _, row in ungrouped.iterrows():
		write_standalone_radar(base, row)
	logger.info("Peer outputs written: %d companies", len(frame))
	return 0


if __name__ == "__main__":
	raise SystemExit(main())
