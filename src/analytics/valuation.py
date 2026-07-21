"""
valuation.py
Valuation analytics module for Nifty 100 companies.

Calculates:
  - FCF yield = (FCF / market_cap_crore) × 100
  - Sector median P/E
  - Valuation flags: Caution (>1.5× median), Discount (<0.7× median), Fair

Outputs:
  - output/valuation_summary.xlsx (all 92 companies)
  - output/valuation_flags.csv (only Caution/Discount companies)
"""

from __future__ import annotations

import logging
import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

logger = logging.getLogger(__name__)

ROOT_DIR = Path(__file__).resolve().parents[2]
DB_PATH = ROOT_DIR / 'data' / 'nifty100.db'
OUTPUT_DIR = ROOT_DIR / 'output'

# Excel styling
HEADER_FILL = PatternFill(fill_type='solid', fgColor='1F2937')
HEADER_FONT = Font(color='FFFFFF', bold=True)
CAUTION_FILL = PatternFill(fill_type='solid', fgColor='FCE4D6')
DISCOUNT_FILL = PatternFill(fill_type='solid', fgColor='E2F0D9')


def load_valuation_data(db_path: Path | None = None) -> pd.DataFrame:
    """
    Load latest year financial ratios, market cap, and sector data.
    """
    path = db_path or DB_PATH
    
    with sqlite3.connect(path) as conn:
        query = """
        SELECT
            fr.company_id,
            c.company_name,
            s.broad_sector,
            fr.free_cash_flow_cr,
            mc.market_cap_crore,
            mc.pe_ratio,
            mc.pb_ratio,
            mc.dividend_yield_pct,
            fr.return_on_equity_pct,
            fr.debt_to_equity,
            fr.composite_quality_score
        FROM financial_ratios fr
        JOIN (
            SELECT company_id, MAX(year) AS yr
            FROM financial_ratios
            GROUP BY company_id
        ) latest ON fr.company_id = latest.company_id AND fr.year = latest.yr
        JOIN companies c ON c.id = fr.company_id
        LEFT JOIN sectors s ON s.company_id = fr.company_id
        LEFT JOIN market_cap mc ON mc.company_id = fr.company_id
            AND mc.year = CAST(SUBSTR(fr.year, 1, 4) AS INTEGER)
        ORDER BY fr.company_id
        """
        df = pd.read_sql_query(query, conn)
    
    return df


def calculate_fcf_yield(df: pd.DataFrame) -> pd.DataFrame:
    """
    Calculate FCF yield = (FCF / market_cap_crore) × 100
    """
    result = df.copy()
    
    result['fcf_yield_pct'] = (
        result['free_cash_flow_cr'] / result['market_cap_crore'] * 100
    ).round(2)
    
    # Replace inf and -inf with NaN
    result['fcf_yield_pct'] = result['fcf_yield_pct'].replace([float('inf'), float('-inf')], pd.NA)
    
    return result


def calculate_sector_median_pe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Calculate sector median P/E for each broad_sector.
    """
    result = df.copy()
    
    # Calculate median P/E per sector
    sector_medians = result.groupby('broad_sector')['pe_ratio'].median().to_dict()
    
    result['sector_median_pe'] = result['broad_sector'].map(sector_medians).round(2)
    
    return result


def assign_valuation_flags(df: pd.DataFrame) -> pd.DataFrame:
    """
    Assign valuation flags based on P/E ratio vs sector median:
      - Caution: P/E > 1.5× sector median
      - Discount: P/E < 0.7× sector median
      - Fair: Otherwise
    """
    result = df.copy()
    
    def _flag(row):
        pe = row['pe_ratio']
        median_pe = row['sector_median_pe']
        
        # Handle missing data
        if pd.isna(pe) or pd.isna(median_pe) or median_pe == 0:
            return 'Fair'
        
        ratio = pe / median_pe
        
        if ratio > 1.5:
            return 'Caution'
        elif ratio < 0.7:
            return 'Discount'
        else:
            return 'Fair'
    
    result['valuation_flag'] = result.apply(_flag, axis=1)
    
    return result


def generate_valuation_summary(
    output_path: Path | None = None,
    db_path: Path | None = None
) -> pd.DataFrame:
    """
    Generate complete valuation summary with all calculations.
    """
    logger.info('Loading valuation data...')
    df = load_valuation_data(db_path)
    
    logger.info(f'Loaded data for {len(df)} companies')
    
    logger.info('Calculating FCF yield...')
    df = calculate_fcf_yield(df)
    
    logger.info('Calculating sector median P/E...')
    df = calculate_sector_median_pe(df)
    
    logger.info('Assigning valuation flags...')
    df = assign_valuation_flags(df)
    
    # Select and order columns for output
    output_cols = [
        'company_id',
        'company_name',
        'broad_sector',
        'pe_ratio',
        'sector_median_pe',
        'pb_ratio',
        'fcf_yield_pct',
        'market_cap_crore',
        'free_cash_flow_cr',
        'return_on_equity_pct',
        'debt_to_equity',
        'composite_quality_score',
        'valuation_flag'
    ]
    
    result = df[output_cols].copy()
    
    # Save to Excel
    path = output_path or (OUTPUT_DIR / 'valuation_summary.xlsx')
    path.parent.mkdir(parents=True, exist_ok=True)
    
    logger.info(f'Writing valuation summary to {path}...')
    result.to_excel(path, index=False, engine='openpyxl')
    
    # Apply styling
    _apply_excel_styling(path)
    
    logger.info(f'✅ Valuation summary written: {len(result)} companies')
    
    return result


def generate_valuation_flags_csv(
    summary_df: pd.DataFrame,
    output_path: Path | None = None
) -> pd.DataFrame:
    """
    Generate CSV with only Caution and Discount flagged companies.
    """
    # Filter to only Caution and Discount
    flagged = summary_df[
        summary_df['valuation_flag'].isin(['Caution', 'Discount'])
    ].copy()
    
    # Select relevant columns
    output_cols = [
        'company_id',
        'company_name',
        'broad_sector',
        'valuation_flag',
        'pe_ratio',
        'sector_median_pe',
        'fcf_yield_pct',
        'market_cap_crore'
    ]
    
    result = flagged[output_cols].copy()
    
    # Sort: Caution first, then by P/E ratio
    result = result.sort_values(
        ['valuation_flag', 'pe_ratio'],
        ascending=[True, False]
    )
    
    # Save to CSV
    path = output_path or (OUTPUT_DIR / 'valuation_flags.csv')
    path.parent.mkdir(parents=True, exist_ok=True)
    
    logger.info(f'Writing valuation flags to {path}...')
    result.to_csv(path, index=False)
    
    logger.info(f'✅ Valuation flags written: {len(result)} flagged companies')
    logger.info(f'   - Caution: {(result["valuation_flag"] == "Caution").sum()}')
    logger.info(f'   - Discount: {(result["valuation_flag"] == "Discount").sum()}')
    
    return result


def _apply_excel_styling(path: Path):
    """
    Apply styling to Excel workbook:
      - Header row: dark background, white bold text
      - Caution rows: orange fill
      - Discount rows: green fill
    """
    wb = load_workbook(path)
    ws = wb.active
    
    # Style header row
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    
    # Find valuation_flag column
    headers = [cell.value for cell in ws[1]]
    if 'valuation_flag' not in headers:
        wb.save(path)
        return
    
    flag_col_idx = headers.index('valuation_flag') + 1
    
    # Style data rows based on flag
    for row_idx in range(2, ws.max_row + 1):
        flag_cell = ws.cell(row=row_idx, column=flag_col_idx)
        flag_value = flag_cell.value
        
        if flag_value == 'Caution':
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = CAUTION_FILL
        elif flag_value == 'Discount':
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = DISCOUNT_FILL
    
    wb.save(path)


def main() -> int:
    """
    Main execution function.
    """
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    logger.info('Starting valuation analysis...')
    
    # Generate summary
    summary_df = generate_valuation_summary()
    
    # Generate flags CSV
    generate_valuation_flags_csv(summary_df)
    
    logger.info('✅ Valuation analysis complete')
    
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
